import pandas as pd
from openpyxl import load_workbook


def dict_to_excel(data: dict, output_path: str):
    '''
    Save structured data dictionary to an Excel file.
    Args:
        data (dict): Structured data to be saved.
        output_path (str): Path to the output Excel file.'''
    df = pd.DataFrame(data, columns=data[0].keys())
    df.to_excel(output_path, index=False)

    # Load workbook and worksheet
    wb = load_workbook(output_path)
    ws = wb.active

    ws["G1"] = "Essentials Total"
    ws["G2"] = '=SUMIF($E$2:$E$1000, "essential", $D$2:$D$1000)'
    ws["H1"] = "Non-Essentials Total"
    ws["H2"] = '=SUMIF($E$2:$E$1000, "", $D$2:$D$1000)'

    wb.save(output_path)
